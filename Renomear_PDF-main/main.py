import os
import re
import shutil
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pdfplumber
from openpyxl import load_workbook


def limpar_nome_arquivo(nome: str) -> str:
    """
    Remove caracteres inválidos para nome de arquivo no Windows/Linux.
    """
    # Remove caracteres proibidos em nomes de arquivo
    nome = re.sub(r'[\\/*?:"<>|]', "", nome)

    # Substitui múltiplos espaços por um único espaço e remove espaços nas pontas
    nome = re.sub(r"\s+", " ", nome).strip()
    return nome


def normalizar_documento(valor: str) -> str:
    """
    Remove qualquer caractere não numérico de CPF/CNPJ.
    """
    # Se o valor for None, retorna string vazia
    if valor is None:
        return ""

    # Mantém apenas dígitos
    return re.sub(r"\D", "", str(valor))


def formatar_documento(documento: str) -> str:
    """
    Formata o documento:
    - CPF: 000.000.000-00
    - CNPJ: mantém como veio, se já estiver formatado
    - CNPJ com 14 dígitos sem máscara: 00.000.000.0000.00
      (seguindo o padrão que aparece nos arquivos)
    """
    # Remove tudo que não for número
    doc = normalizar_documento(documento)

    # Se tiver 11 dígitos, assume CPF
    if len(doc) == 11:
        return f"{doc[0:3]}.{doc[3:6]}.{doc[6:9]}-{doc[9:11]}"

    # Se tiver 14 dígitos, assume CNPJ no formato adotado pelo projeto
    if len(doc) == 14:
        return f"{doc[0:2]}.{doc[2:5]}.{doc[5:8]}.{doc[8:12]}.{doc[12:14]}"

    # Se não bater com CPF/CNPJ esperados, devolve como veio
    return documento.strip()


def carregar_uens_do_excel(caminho_excel: str) -> dict:
    """
    Lê a planilha e cria um mapa:
    CPF/CNPJ -> UEN

    Estrutura esperada:
    Coluna A = CPF/CNPJ
    Coluna C = UEN
    """
    # Carrega a planilha Excel usando os valores calculados das células
    wb = load_workbook(caminho_excel, data_only=True)
    ws = wb.active

    # Dicionário final: documento normalizado -> UEN
    mapa_uens = {}

    # Começa da linha 2, presumindo cabeçalho na linha 1
    for row in range(2, ws.max_row + 1):
        documento = ws.cell(row=row, column=1).value
        uen = ws.cell(row=row, column=3).value

        # Só adiciona se houver documento e UEN preenchidos
        if documento and uen is not None:
            doc_normalizado = normalizar_documento(documento)
            mapa_uens[doc_normalizado] = str(uen).strip()

    return mapa_uens


def extrair_texto_pdf(pdf_path: str) -> str:
    """
    Extrai todo o texto do PDF.
    """
    texto_completo = []

    try:
        # Abre o PDF e percorre página por página
        with pdfplumber.open(pdf_path) as pdf:
            for pagina in pdf.pages:
                texto = pagina.extract_text()

                # Só adiciona se a página tiver texto extraído
                if texto:
                    texto_completo.append(texto)
    except Exception as e:
        # Relança erro com mensagem mais amigável
        raise RuntimeError(f"Erro ao abrir o PDF: {e}")

    # Junta o texto de todas as páginas
    texto = "\n".join(texto_completo)

    # Normaliza quebras de linha
    texto = texto.replace("\r", "\n")
    texto = re.sub(r"\n+", "\n", texto)
    return texto


def extrair_do_nome_arquivo(nome_arquivo: str):
    """
    Extrai documento e nome a partir do nome do arquivo.

    Exemplos aceitos:
    - 006.369.721-19 - MARIA ISABEL.pdf
    - 10.787.957.0001.83 IEG LTDA.pdf
    - 11.158.756.0001.80 CRIMET ... EIRELI - ME.pdf
    """
    # Remove caminho e extensão, ficando apenas com o nome base
    nome_base = os.path.splitext(os.path.basename(nome_arquivo))[0].strip()

    # Lista de padrões aceitos para identificar documento e nome
    padroes = [
        # CPF no começo com separador " - "
        r"^(\d{3}\.\d{3}\.\d{3}-\d{2})\s*-\s*(.+)$",

        # CNPJ no padrão do projeto: 10.787.957.0001.83 NOME
        r"^(\d{2}\.\d{3}\.\d{3}\.\d{4}\.\d{2})\s+(.+)$",

        # CNPJ no formato mais comum: 10.787.957/0001-83 NOME
        r"^(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})\s+(.+)$",
    ]

    # Testa cada padrão até encontrar correspondência
    for padrao in padroes:
        match = re.match(padrao, nome_base, re.IGNORECASE)
        if match:
            documento = match.group(1).strip()
            nome = match.group(2).strip()

            # Normaliza espaços no nome
            nome = re.sub(r"\s+", " ", nome)
            return documento, nome

    # Se nada for encontrado, retorna vazio
    return None, None


def extrair_documento_e_nome_do_texto(texto: str):
    """
    Tenta extrair CPF/CNPJ e Nome/Razão Social do conteúdo do PDF.
    """
    # Separa o texto em linhas não vazias
    linhas = [linha.strip() for linha in texto.split("\n") if linha.strip()]

    # 1) Tentativa por regex em bloco
    # Procura padrões mais estruturados no texto inteiro
    padroes_bloco = [
        r"CPF\s+Nome Completo\s*\n\s*"
        r"(\d{3}\.\d{3}\.\d{3}-\d{2})\s+"
        r"([A-ZÁÀÂÃÉÈÊÍÌÎÓÒÔÕÚÙÛÇa-záàâãéèêíìîóòôõúùûç0-9 ,.&'()/\-]+)",

        r"CNPJ\s+(?:Raz[aã]o Social|Nome Empresarial)\s*\n\s*"
        r"(\d{2}\.\d{3}\.\d{3}(?:[./]\d{4}[.-]\d{2}|\.\d{4}\.\d{2}))\s+"
        r"([A-ZÁÀÂÃÉÈÊÍÌÎÓÒÔÕÚÙÛÇa-záàâãéèêíìîóòôõúùûç0-9 ,.&'()/\-]+)",

        r"(CPF|CNPJ)\s+(?:Nome Completo|Raz[aã]o Social|Nome Empresarial)\s*\n\s*"
        r"(\d{3}\.\d{3}\.\d{3}-\d{2}|\d{2}\.\d{3}\.\d{3}(?:[./]\d{4}[.-]\d{2}|\.\d{4}\.\d{2}))\s+"
        r"([A-ZÁÀÂÃÉÈÊÍÌÎÓÒÔÕÚÙÛÇa-záàâãéèêíìîóòôõúùûç0-9 ,.&'()/\-]+)",
    ]

    for padrao in padroes_bloco:
        match = re.search(padrao, texto, re.IGNORECASE)
        if match:
            # Alguns padrões capturam 2 grupos, outros 3
            if len(match.groups()) == 2:
                documento = match.group(1).strip()
                nome = match.group(2).strip()
            else:
                documento = match.group(2).strip()
                nome = match.group(3).strip()

            # Corta o nome caso comece a "invadir" outros campos do PDF
            nome = re.split(
                r"\n|Natureza do Rendimento|3\. Rendimentos|Rendimentos Tribut[aá]veis|Fonte Pagadora|Benefici[aá]rio",
                nome,
                maxsplit=1,
                flags=re.IGNORECASE
            )[0].strip()

            # Normaliza espaços
            nome = re.sub(r"\s+", " ", nome)
            return documento, nome

    # 2) Tentativa linha a linha
    for i, linha in enumerate(linhas):
        linha_upper = linha.upper()

        # Caso Pessoa Física
        if "CPF" in linha_upper and "NOME COMPLETO" in linha_upper:
            # Procura nas próximas linhas o padrão: CPF + nome
            for j in range(i + 1, min(i + 4, len(linhas))):
                match = re.search(
                    r"(\d{3}\.\d{3}\.\d{3}-\d{2})\s+(.+)",
                    linhas[j],
                    re.IGNORECASE
                )
                if match:
                    documento = match.group(1).strip()
                    nome = match.group(2).strip()

                    # Remove conteúdo adicional após o nome
                    nome = re.split(
                        r"Natureza do Rendimento|3\. Rendimentos|Rendimentos Tribut[aá]veis|Fonte Pagadora|Benefici[aá]rio",
                        nome,
                        maxsplit=1,
                        flags=re.IGNORECASE
                    )[0].strip()

                    nome = re.sub(r"\s+", " ", nome)
                    return documento, nome

        # Caso Pessoa Jurídica
        if "CNPJ" in linha_upper and ("RAZ" in linha_upper or "EMPRESARIAL" in linha_upper):
            # Procura nas próximas linhas o padrão: CNPJ + razão social
            for j in range(i + 1, min(i + 4, len(linhas))):
                match = re.search(
                    r"(\d{2}\.\d{3}\.\d{3}(?:[./]\d{4}[.-]\d{2}|\.\d{4}\.\d{2}))\s+(.+)",
                    linhas[j],
                    re.IGNORECASE
                )
                if match:
                    documento = match.group(1).strip()
                    nome = match.group(2).strip()

                    # Remove conteúdo adicional após o nome
                    nome = re.split(
                        r"Natureza do Rendimento|3\. Rendimentos|Rendimentos Tribut[aá]veis|Fonte Pagadora|Benefici[aá]rio",
                        nome,
                        maxsplit=1,
                        flags=re.IGNORECASE
                    )[0].strip()

                    nome = re.sub(r"\s+", " ", nome)
                    return documento, nome

    # 3) Busca genérica em qualquer linha com documento + nome
    for linha in linhas:
        # Tenta achar CPF seguido de nome
        match_cpf = re.search(r"(\d{3}\.\d{3}\.\d{3}-\d{2})\s+(.+)", linha)
        if match_cpf:
            documento = match_cpf.group(1).strip()
            nome = match_cpf.group(2).strip()

            # Só aceita nomes com tamanho mínimo
            if len(nome) >= 3:
                return documento, re.sub(r"\s+", " ", nome)

        # Tenta achar CNPJ seguido de nome
        match_cnpj = re.search(
            r"(\d{2}\.\d{3}\.\d{3}(?:[./]\d{4}[.-]\d{2}|\.\d{4}\.\d{2}))\s+(.+)",
            linha
        )
        if match_cnpj:
            documento = match_cnpj.group(1).strip()
            nome = match_cnpj.group(2).strip()

            # Só aceita nomes com tamanho mínimo
            if len(nome) >= 3:
                return documento, re.sub(r"\s+", " ", nome)

    # Nada encontrado
    return None, None


def extrair_documento_e_nome(pdf_path: str):
    """
    Estratégia final:
    1. Tenta extrair do conteúdo do PDF
    2. Se falhar, tenta extrair do nome do arquivo
    """
    # Primeiro extrai o texto do PDF
    texto = extrair_texto_pdf(pdf_path)

    # Tenta identificar documento e nome pelo conteúdo
    documento, nome = extrair_documento_e_nome_do_texto(texto)
    if documento and nome:
        return formatar_documento(documento), nome

    # Se falhar, tenta usar o nome do arquivo
    documento, nome = extrair_do_nome_arquivo(pdf_path)
    if documento and nome:
        return formatar_documento(documento), nome

    return None, None


def montar_nome_arquivo(documento: str, nome: str, uen: str = "") -> str:
    """
    Monta o nome final do arquivo.
    """
    # Se houver UEN, inclui no final do nome
    if uen:
        nome_final = f"{documento} - {nome} - {uen}.pdf"
    else:
        nome_final = f"{documento} - {nome}.pdf"

    # Garante que o nome final seja válido para o sistema operacional
    return limpar_nome_arquivo(nome_final)


class AppRenomeadorPDF:
    def __init__(self, root):
        self.root = root
        self.root.title("Renomeador de PDFs por Documento e Nome")
        self.root.geometry("780x540")
        self.root.minsize(720, 480)

        # Variáveis ligadas aos campos da interface
        self.input_folder = tk.StringVar()
        self.output_folder = tk.StringVar()
        self.excel_file = tk.StringVar()
        self.prefixo = tk.StringVar(value="")

        self.criar_interface()

    def criar_interface(self):
        # Frame principal
        frame = ttk.Frame(self.root, padding=15)
        frame.pack(fill="both", expand=True)

        # Título principal
        titulo = ttk.Label(
            frame,
            text="Renomear PDFs com CPF/CNPJ e Nome",
            font=("Segoe UI", 14, "bold")
        )
        titulo.pack(anchor="w", pady=(0, 10))

        # Texto explicativo
        subtitulo = ttk.Label(
            frame,
            text=(
                "Selecione a pasta de origem e a pasta de destino. "
                "O sistema tenta extrair CPF/CNPJ e Nome/Razão Social do PDF. "
                "Se não conseguir, tenta usar o próprio nome do arquivo. "
                "A UEN será buscada no Excel, se informado."
            ),
            wraplength=730,
            justify="left"
        )
        subtitulo.pack(anchor="w", pady=(0, 15))

        # Campo da pasta de origem
        origem_frame = ttk.LabelFrame(frame, text="Pasta de origem", padding=10)
        origem_frame.pack(fill="x", pady=5)

        ttk.Entry(origem_frame, textvariable=self.input_folder).pack(
            side="left", fill="x", expand=True, padx=(0, 10)
        )
        ttk.Button(origem_frame, text="Selecionar", command=self.selecionar_origem).pack(side="left")

        # Campo da pasta de destino
        destino_frame = ttk.LabelFrame(frame, text="Pasta de destino", padding=10)
        destino_frame.pack(fill="x", pady=5)

        ttk.Entry(destino_frame, textvariable=self.output_folder).pack(
            side="left", fill="x", expand=True, padx=(0, 10)
        )
        ttk.Button(destino_frame, text="Selecionar", command=self.selecionar_destino).pack(side="left")

        # Campo do Excel opcional
        excel_frame = ttk.LabelFrame(frame, text="Arquivo Excel (opcional para UEN)", padding=10)
        excel_frame.pack(fill="x", pady=5)

        ttk.Entry(excel_frame, textvariable=self.excel_file).pack(
            side="left", fill="x", expand=True, padx=(0, 10)
        )
        ttk.Button(excel_frame, text="Selecionar", command=self.selecionar_xlsx).pack(side="left")

        # Campo para prefixo opcional
        prefixo_frame = ttk.LabelFrame(frame, text="Opções", padding=10)
        prefixo_frame.pack(fill="x", pady=5)

        ttk.Label(prefixo_frame, text="Prefixo opcional no nome do arquivo:").pack(side="left", padx=(0, 10))
        ttk.Entry(prefixo_frame, textvariable=self.prefixo, width=30).pack(side="left")

        # Botões de ação
        botoes = ttk.Frame(frame)
        botoes.pack(fill="x", pady=15)

        self.btn_processar = ttk.Button(
            botoes,
            text="Processar PDFs",
            command=self.iniciar_processamento
        )
        self.btn_processar.pack(side="left")

        ttk.Button(
            botoes,
            text="Limpar log",
            command=self.limpar_log
        ).pack(side="left", padx=10)

        # Barra de progresso
        self.progress = ttk.Progressbar(frame, mode="determinate")
        self.progress.pack(fill="x", pady=(0, 10))

        # Área de log
        log_frame = ttk.LabelFrame(frame, text="Log de processamento", padding=10)
        log_frame.pack(fill="both", expand=True)

        self.log_text = tk.Text(log_frame, height=15, wrap="word")
        self.log_text.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_text.configure(yscrollcommand=scrollbar.set)

    def selecionar_origem(self):
        # Abre seletor de diretório para pasta de origem
        pasta = filedialog.askdirectory(title="Selecione a pasta com os PDFs")
        if pasta:
            self.input_folder.set(pasta)

    def selecionar_destino(self):
        # Abre seletor de diretório para pasta de destino
        pasta = filedialog.askdirectory(title="Selecione a pasta de destino")
        if pasta:
            self.output_folder.set(pasta)

    def selecionar_xlsx(self):
        # Abre seletor de arquivo Excel
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
        )
        if arquivo:
            self.excel_file.set(arquivo)

    def log(self, mensagem: str):
        # Escreve uma mensagem no log da interface
        self.log_text.insert("end", mensagem + "\n")
        self.log_text.see("end")
        self.root.update_idletasks()

    def limpar_log(self):
        # Limpa toda a área de log
        self.log_text.delete("1.0", "end")

    def iniciar_processamento(self):
        # Obtém os valores preenchidos na interface
        origem = self.input_folder.get().strip()
        destino = self.output_folder.get().strip()
        caminho_excel = self.excel_file.get().strip()

        # Validações básicas
        if not origem:
            messagebox.showwarning("Atenção", "Selecione a pasta de origem.")
            return

        if not destino:
            messagebox.showwarning("Atenção", "Selecione a pasta de destino.")
            return

        if not os.path.isdir(origem):
            messagebox.showerror("Erro", "A pasta de origem é inválida.")
            return

        if not os.path.isdir(destino):
            messagebox.showerror("Erro", "A pasta de destino é inválida.")
            return

        if caminho_excel and not os.path.isfile(caminho_excel):
            messagebox.showerror("Erro", "O arquivo Excel informado é inválido.")
            return

        # Desabilita botão para evitar múltiplos processamentos simultâneos
        self.btn_processar.config(state="disabled")
        self.progress["value"] = 0
        self.limpar_log()

        # Executa o processamento em uma thread separada para não travar a interface
        thread = threading.Thread(target=self.processar_pdfs, daemon=True)
        thread.start()

    def processar_pdfs(self):
        # Lê os parâmetros atuais da interface
        origem = self.input_folder.get().strip()
        destino = self.output_folder.get().strip()
        prefixo = self.prefixo.get().strip()
        caminho_excel = self.excel_file.get().strip()

        # Carrega o mapa de UENs, se houver Excel informado
        mapa_uens = {}
        if caminho_excel:
            try:
                mapa_uens = carregar_uens_do_excel(caminho_excel)
                self.log(f"[INFO] Excel carregado com {len(mapa_uens)} documento(s).")
            except Exception as e:
                self.log(f"[ERRO] Falha ao carregar Excel: {e}")
                self.btn_processar.config(state="normal")
                return

        # Lista apenas arquivos PDF da pasta de origem
        arquivos_pdf = [f for f in os.listdir(origem) if f.lower().endswith(".pdf")]

        if not arquivos_pdf:
            self.log("Nenhum PDF encontrado na pasta de origem.")
            self.btn_processar.config(state="normal")
            return

        total = len(arquivos_pdf)
        self.progress["maximum"] = total

        processados = 0
        erros = 0

        # Processa cada PDF individualmente
        for i, arquivo in enumerate(arquivos_pdf, start=1):
            caminho_pdf = os.path.join(origem, arquivo)

            try:
                # Extrai documento e nome do PDF ou do nome do arquivo
                documento, nome = extrair_documento_e_nome(caminho_pdf)

                if not documento or not nome:
                    self.log(f"[ERRO] Não foi possível extrair CPF/CNPJ e Nome: {arquivo}")
                    erros += 1
                    self.progress["value"] = i
                    continue

                # Normaliza o documento para consultar no Excel
                doc_normalizado = normalizar_documento(documento)
                uen = mapa_uens.get(doc_normalizado, "")

                # Loga o resultado da busca de UEN, se Excel foi informado
                if caminho_excel:
                    if uen:
                        self.log(f"[INFO] UEN encontrada para {documento}: {uen}")
                    else:
                        self.log(f"[INFO] Nenhuma UEN encontrada para {documento}")

                # Monta o novo nome do arquivo
                novo_nome = montar_nome_arquivo(documento, nome, uen)

                # Adiciona prefixo, se configurado
                if prefixo:
                    novo_nome = limpar_nome_arquivo(f"{prefixo} {novo_nome}")

                destino_final = os.path.join(destino, novo_nome)

                # Evita sobrescrever arquivos existentes
                contador = 1
                base_nome, ext = os.path.splitext(destino_final)
                while os.path.exists(destino_final):
                    destino_final = f"{base_nome} ({contador}){ext}"
                    contador += 1

                # Copia o arquivo renomeado para a pasta de destino
                shutil.copy2(caminho_pdf, destino_final)
                self.log(f"[OK] {arquivo}  ->  {os.path.basename(destino_final)}")
                processados += 1

            except Exception as e:
                # Captura erros inesperados por arquivo
                self.log(f"[ERRO] {arquivo}: {e}")
                erros += 1

            # Atualiza barra de progresso
            self.progress["value"] = i

        # Reabilita o botão ao final do processamento
        self.btn_processar.config(state="normal")

        # Mostra resumo final
        messagebox.showinfo(
            "Concluído",
            f"Processamento finalizado.\n\n"
            f"Sucesso: {processados}\n"
            f"Erros: {erros}\n"
            f"Total: {total}"
        )


if __name__ == "__main__":
    # Cria a janela principal e inicia a aplicação
    root = tk.Tk()
    app = AppRenomeadorPDF(root)
    root.mainloop()